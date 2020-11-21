import csv
import openpyxl
import wx
import wx

class MainWindow(wx.Frame):
    def __init__(self, parent, title):
        wx.Frame.__init__(self, parent, title = title, size=(500, 200))
        self.panel = wx.Panel(self)
        row1 = wx.StaticBoxSizer(wx.VERTICAL, self.panel, 'Selezionare il file Export:')
        self.fileCtrl = wx.FilePickerCtrl(self.panel, message="Select file", style=wx.FLP_USE_TEXTCTRL, size=(290, 25))
        row1.Add(self.fileCtrl, 0, wx.ALL, 10)
        self.panel.SetSizer(row1)
        self.Show()

app = wx.App(False)
win = MainWindow(None, "File selector")
app.MainLoop()

from openpyxl import load_workbook
wb = load_workbook(r"base.xlsx", data_only=True)
sh = wb.active
sh.cell(row=1, column=6).value = "ParentOC"
i = 0
for figa in sh.iter_rows(min_row=2,max_col=1,max_row=sh.max_row,values_only=True):
    print(figa)
    parent_id = None 
    i+=1
    with open('export.csv', mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)

        line_count = 0
        for row in csv_reader:
            if(row["Sku"] == figa[0] ): 
                parent_id = row["Parent Product ID"]
                break

    with open('export.csv', mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row2 in csv_reader:
            if(row2["ID"] == parent_id and row2["Parent Product ID"] == "0"):
                print(row2["Sku"])
                sh.cell(row=i+1, column=6).value = row2["Sku"]
                break

    
wb.save(r"base_edited.xlsx")